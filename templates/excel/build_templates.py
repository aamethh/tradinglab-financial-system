"""Genera las 7 plantillas Excel de AmethQuant en templates/excel/.

Ejecutar:  python templates/excel/build_templates.py
Cada plantilla implementa métodos oficiales del Method Inventory (docs/method-inventory.md).
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent
NAVY = "1F2A44"
BLUE = "2563EB"
GREY = "F1F5F9"


def style_sheet(ws, headers, widths, title, subtitle):
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True, color=NAVY)
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=9, italic=True, color="64748B")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]
    ws.freeze_panes = "A5"


def fill_rows(ws, rows, start=5):
    for r, row in enumerate(rows, start):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=GREY)
            cell.alignment = Alignment(wrap_text=True, vertical="top")


# 1 ── Forensic checklist (Schilit Framework — método oficial #1)
def forensic_checklist():
    wb = Workbook(); ws = wb.active; ws.title = "Checklist Schilit"
    style_sheet(ws, ["#", "Bandera", "Umbral / regla", "Peso", "¿Activa? (SI/NO)", "Evidencia (cita del filing)", "Puntos"],
                [4, 38, 42, 8, 14, 45, 10],
                "AmethQuant — Forensic Checklist (Schilit Framework)",
                "Score 0-100 interno (Proposed Addition P1). El entregable cliente es la tabla de banderas con evidencia.")
    flags = [
        ("FCO / Utilidad Neta", "< 1.0x", 15), ("FCO negativo con UN positiva", "FCO<0 y UN>0", 12),
        ("DSO", "> 1.5x histórico", 12), ("One-time gains recurrentes", "2+ períodos consecutivos", 10),
        ("Capitalización agresiva", "Capex/Ingresos >1.5x hist. o intangibles +30% YoY", 8),
        ("Patrimonio operativo", "Negativo sin revaluaciones", 10),
        ("Asunto de Énfasis del auditor", "Presente en informe", 10),
        ("Cobertura de intereses", "< 2.0x", 8), ("ROIC vs WACC", "ROIC < WACC", 6),
        ("Partes relacionadas", "Material sin detalle", 4),
        ("Divergencia vs sector", "Crecimiento sin explicación", 3),
        ("Deuda/EBITDA", "En alza 3+ períodos", 2),
    ]
    for i, (flag, rule, w) in enumerate(flags, 1):
        r = 4 + i
        fill_rows(ws, [[i, flag, rule, w, "NO", ""]], start=r)
        ws.cell(row=r, column=7, value=f'=IF(E{r}="SI",D{r},0)')
    ws["F18"] = "FORENSIC SCORE:"; ws["F18"].font = Font(bold=True)
    ws["G18"] = "=SUM(G5:G16)"; ws["G18"].font = Font(bold=True, color="DC2626")
    ws["F19"] = "Nivel:"; ws["G19"] = '=IF(G18>=60,"CRITICO",IF(G18>=35,"ALTO",IF(G18>=15,"MODERADO","BAJO")))'
    wb.save(OUT / "01_forensic_checklist.xlsx")


# 2 ── Industrial company analysis (método oficial #9: motor de ratios)
def industrial():
    wb = Workbook(); ws = wb.active; ws.title = "Inputs"
    style_sheet(ws, ["Line item", "FY-2", "FY-1", "FY actual", "Fuente (filing + página)"],
                [34, 14, 14, 14, 40],
                "AmethQuant — Análisis Empresa Industrial",
                "Datos primarios obligatorios. No usar agregadores sin verificar contra el filing.")
    items = ["Ingresos", "Costo de ventas", "EBITDA", "Utilidad neta", "FCO", "Capex",
             "Cuentas por cobrar", "Inventario", "Activos totales", "Deuda financiera",
             "Patrimonio", "Gastos por intereses", "Efectivo"]
    fill_rows(ws, [[i, None, None, None, ""] for i in items])
    ws2 = wb.create_sheet("Ratios")
    style_sheet(ws2, ["Ratio", "Fórmula", "FY actual", "Benchmark", "Señal"],
                [28, 38, 14, 18, 24],
                "Ratios (15+) — perfil completo",
                "Profitabilidad, apalancamiento, liquidez, eficiencia, calidad de caja.")
    ratios = [
        ("Margen EBITDA", "=Inputs!D7/Inputs!D5 (EBITDA/Ingresos)"), ("Margen neto", "UN/Ingresos"),
        ("ROE", "UN/Patrimonio"), ("ROA", "UN/Activos"), ("ROIC", "NOPAT/Capital invertido"),
        ("FCO/UN  ⚠ Schilit", "FCO/Utilidad neta — umbral 1.0x"), ("DSO  ⚠ Schilit", "CxC/Ingresos × 365"),
        ("Deuda/EBITDA", "Deuda financiera/EBITDA"), ("Cobertura intereses", "EBITDA/Intereses — umbral 2.0x"),
        ("Current ratio", "Activo corriente/Pasivo corriente"), ("Capex/Ingresos", "vs promedio histórico"),
        ("Crecimiento ingresos YoY", "vs sector"), ("D/E", "Deuda/Patrimonio"),
        ("Margen bruto", "(Ingresos-CoGS)/Ingresos"), ("CCC", "DSO+DIO-DPO"),
    ]
    fill_rows(ws2, [[a, b, None, "", ""] for a, b in ratios])
    wb.save(OUT / "02_industrial_analysis.xlsx")


# 3 ── Bank analysis (métodos oficiales #12-15 + marco bancario P2)
def bank():
    wb = Workbook(); ws = wb.active; ws.title = "Marco Bancario"
    style_sheet(ws, ["Dimensión", "Métrica", "Valor", "Benchmark regional", "Fuente", "Señal 🟢🟡🔴"],
                [22, 34, 14, 22, 30, 14],
                "AmethQuant — Bank Analysis (marco FIG)",
                "NUNCA analizar un banco como industrial. Caso de referencia: BGFG (OUTPERFORM).")
    rows = [
        ("Loan book", "Cartera bruta y mix por segmento", None, "", "", ""),
        ("Loan book", "Crecimiento cartera YoY", None, "", "", ""),
        ("Calidad", "NPL ratio (morosidad)", None, "Panamá: comparar SBP", "", ""),
        ("Calidad", "Cobertura de provisiones / NPL", None, "BGFG: 124.3% (↓ de 152.6%)", "", ""),
        ("Rentabilidad", "ROA", None, "", "", ""),
        ("Rentabilidad", "ROAE", None, "BGFG: 21.2% top-quartile", "", ""),
        ("Rentabilidad", "NIM (margen de interés neto)", None, "", "", ""),
        ("Eficiencia", "Cost-to-income", None, "BGFG: 28.1% best-in-class", "", ""),
        ("Capital", "Ratio de adecuación (BIS III)", None, "Mínimo Panamá 8-10.5%", "", ""),
        ("Liquidez", "Liquidez legal / LCR", None, "Mínimo SBP 30%", "", ""),
        ("Fondeo", "Depósitos / pasivos totales", None, "", "", ""),
        ("Fondeo", "Crecimiento de depósitos YoY", None, "", "", ""),
        ("Concentración", "Top 20 deudores / cartera", None, "", "", ""),
        ("Concentración", "Exposición geográfica", None, "BGFG: Panamá ~85.8%", "", ""),
        ("Soberano", "Exposición a deuda soberana / activos", None, "", "", ""),
        ("ROE Decomposition", "Margen × rotación × apalancamiento", None, "", "", ""),
        ("Valoración", "P/B vs peers regionales", None, "", "", ""),
        ("AUM (si aplica)", "AUM y crecimiento vs loans", None, "BGFG: AUM 3× más rápido", "", ""),
        ("Regulatorio", "Cambios normativos / sanciones", None, "", "", ""),
        ("Reputacional", "Litigios, prensa, GAFI/listas", None, "", "", ""),
    ]
    fill_rows(ws, rows)
    wb.save(OUT / "03_bank_analysis.xlsx")


# 4 ── Credit risk analysis (métodos oficiales #16-17)
def credit():
    wb = Workbook(); ws = wb.active; ws.title = "Credit Risk"
    style_sheet(ws, ["Dimensión", "Métrica", "Valor", "Umbral / referencia", "Fuente"],
                [24, 36, 16, 28, 30],
                "AmethQuant — Credit / Solvency Review",
                "Marco: duration analysis, credit risk assessment, peer benchmarking, macro sensitivity (FGIN).")
    rows = [
        ("Servicio de deuda", "Cobertura de intereses (EBITDA/Int.)", None, "< 2.0x = estrés", ""),
        ("Servicio de deuda", "DSCR", None, "< 1.2x = alerta", ""),
        ("Apalancamiento", "Deuda/EBITDA", None, "> 4x = alto", ""),
        ("Apalancamiento", "Deuda neta/Patrimonio", None, "", ""),
        ("Liquidez", "Current ratio", None, "< 1.0 = alerta", ""),
        ("Liquidez", "Caja / deuda corto plazo", None, "", ""),
        ("Calidad de caja", "FCO/UN  ⚠ Schilit", None, "< 1.0x = utilidades sin caja", ""),
        ("Perfil crediticio", "Rating dominante de cartera/emisor", None, "FGIN: BBB-/BB+", ""),
        ("Tasa", "Duración promedio", None, "FGIN: 2.1 años", ""),
        ("Vencimientos", "Muro de vencimientos 24 meses", None, "", ""),
        ("Macro", "Sensibilidad a tasas/ciclo", None, "", ""),
        ("Peers", "Spread vs comparables", None, "", ""),
    ]
    fill_rows(ws, rows)
    wb.save(OUT / "04_credit_risk.xlsx")


# 5 ── Valuation scenarios (métodos oficiales #6-7: DCF 3 escenarios + WACC×TGR)
def valuation():
    wb = Workbook(); ws = wb.active; ws.title = "DCF 3 Escenarios"
    style_sheet(ws, ["Parámetro", "Bear", "Base", "Bull", "Nota"],
                [26, 14, 14, 14, 40],
                "AmethQuant — Valuation Scenarios (DCF 3 escenarios)",
                "Blend 50/50 Gordon Growth + Múltiplo de Salida. Threshold: BUY >15% upside · HOLD 0-15% · SELL <0%.")
    params = [
        ("CAGR Ingresos", "Referencia MSFT: 7.8% / 13.4% / 16.0%"),
        ("Margen EBITDA", "MSFT: 49% / 52% / 55%"),
        ("Capex % ingresos", ""), ("Tax rate", ""),
        ("WACC (CAPM)", "Rf + Beta × ERP — MSFT base 9.2%"),
        ("TGR (crecimiento terminal)", "MSFT: 2.5% / 3.5% / 4.0%"),
        ("Valor Gordon Growth", ""), ("Valor Múltiplo Salida", ""),
        ("Valor intrínseco (50/50)", "=PROMEDIO(Gordon, Múltiplo)"),
        ("Precio actual", ""), ("Upside %", "=(Intrínseco/Precio)-1"),
        ("Rating", '=SI(upside>15%,"BUY",SI(upside>=0,"HOLD","SELL/AVOID"))'),
    ]
    fill_rows(ws, [[p, None, None, None, n] for p, n in params])
    ws2 = wb.create_sheet("Sensibilidad WACCxTGR")
    ws2["A1"] = "Tabla de sensibilidad 9×5 — WACC (filas) × TGR (columnas)"
    ws2["A1"].font = Font(bold=True, color=NAVY)
    ws2["A3"] = "WACC \\ TGR"
    tgrs = [2.0, 2.5, 3.0, 3.5, 4.0]
    waccs = [7.0, 7.5, 8.0, 8.5, 9.0, 9.5, 10.0, 10.5, 11.0]
    for j, t in enumerate(tgrs, 2):
        c = ws2.cell(row=3, column=j, value=t / 100); c.font = Font(bold=True); c.number_format = "0.0%"
    for i, w in enumerate(waccs, 4):
        c = ws2.cell(row=i, column=1, value=w / 100); c.font = Font(bold=True); c.number_format = "0.0%"
    ws2["H3"] = "Llenar con valor intrínseco por combinación (output de core/dcf.py)"
    wb.save(OUT / "05_valuation_scenarios.xlsx")


# 6 ── Job application tracker
def jobs():
    wb = Workbook(); ws = wb.active; ws.title = "Tracker"
    style_sheet(ws, ["Fecha", "Rol", "Empresa", "Fuente", "URL", "Fit (0-100)", "Estado",
                     "Mensaje enviado", "Próximo paso", "Notas"],
                [12, 28, 22, 14, 30, 11, 14, 30, 22, 26],
                "AmethQuant — Job Application Tracker",
                "Estados: identificado → aplicado → entrevista → oferta/rechazado. Nada se envía sin aprobación.")
    fill_rows(ws, [["", "", "", "", "", None, "identificado", "", "", ""] for _ in range(15)])
    wb.save(OUT / "06_job_tracker.xlsx")


# 7 ── Client pipeline tracker
def pipeline():
    wb = Workbook(); ws = wb.active; ws.title = "Pipeline"
    style_sheet(ws, ["Fecha", "Lead", "Organización", "Canal", "Servicio de interés",
                     "Etapa", "Valor estimado $", "Próxima acción", "Fecha acción", "Notas"],
                [12, 22, 22, 12, 26, 14, 14, 24, 12, 26],
                "AmethQuant — Client Pipeline",
                "Etapas: nuevo → contactado → reunión → propuesta → ganado/perdido. Servicios y precios: docs/services-pricing.md.")
    fill_rows(ws, [["", "", "", "", "", "nuevo", None, "", "", ""] for _ in range(15)])
    wb.save(OUT / "07_client_pipeline.xlsx")


if __name__ == "__main__":
    for fn in (forensic_checklist, industrial, bank, credit, valuation, jobs, pipeline):
        fn()
        print("OK", fn.__name__)
